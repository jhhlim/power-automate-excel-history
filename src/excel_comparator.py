"""Core Excel file comparison logic."""

import openpyxl
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Any
import pandas as pd


class ExcelComparator:
    """Compares two Excel files and detects changes."""

    def __init__(self):
        pass

    def load_excel(self, file_path: str) -> Dict[str, List[List[Any]]]:
        """
        Load Excel file and convert sheets to dictionaries with data.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary with sheet names as keys and data as values
        """
        workbook = load_workbook(file_path, data_only=True)
        sheets_data = {}
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            data = []
            
            for row in worksheet.iter_rows(values_only=True):
                data.append(list(row))
            
            sheets_data[sheet_name] = data
        
        return sheets_data

    def compare_sheets(self, old_data: List[List[Any]], new_data: List[List[Any]]) -> Dict[str, Any]:
        """
        Compare two sheets and identify differences.
        
        Args:
            old_data: Previous version of sheet data
            new_data: Current version of sheet data
            
        Returns:
            Dictionary containing detected changes
        """
        changes = {
            'rows_added': [],
            'rows_deleted': [],
            'rows_modified': [],
            'headers_changed': False,
            'summary': {}
        }
        
        # Convert to sets of tuples for comparison
        old_set = set(tuple(row) if row else tuple() for row in old_data)
        new_set = set(tuple(row) if row else tuple() for row in new_data)
        
        # Find added and deleted rows
        added = new_set - old_set
        deleted = old_set - new_set
        
        changes['rows_added'] = [list(row) for row in added]
        changes['rows_deleted'] = [list(row) for row in deleted]
        changes['rows_modified'] = len(added) + len(deleted)
        
        # Check if headers changed (first row)
        if old_data and new_data:
            if old_data[0] != new_data[0]:
                changes['headers_changed'] = True
        
        return changes

    def compare_files(self, old_file: str, new_file: str) -> Dict[str, Dict[str, Any]]:
        """
        Compare two Excel files.
        
        Args:
            old_file: Path to previous version
            new_file: Path to current version
            
        Returns:
            Dictionary with changes per sheet
        """
        old_data = self.load_excel(old_file)
        new_data = self.load_excel(new_file)
        
        all_changes = {}
        
        # Compare all sheets in new file
        for sheet_name in new_data.keys():
            old_sheet = old_data.get(sheet_name, [])
            new_sheet = new_data.get(sheet_name, [])
            
            changes = self.compare_sheets(old_sheet, new_sheet)
            
            # Only include sheets with changes
            if changes['rows_added'] or changes['rows_deleted'] or changes['headers_changed']:
                all_changes[sheet_name] = changes
        
        return all_changes
